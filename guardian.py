"""
Author : Vishak Raj
E-mail ID : vishak.shanmu@gmail.com

Gives and save the news from the guardian news paper
And the guardian have limitaiton, to get full access contact the guardian 

"""

import json
import requests
import html2text,pandas as pd
import csv,xlsxwriter
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import re
import schedule
import time
import xlwt

# Refer the doc - https://open-platform.theguardian.com/documentation/

def guardian():
    API_KEY = ""  #get your api-key from https://open-platform.theguardian.com/access/
    
    API_ENDPOINT = 'http://content.guardianapis.com/search' # api-endpoit means calling code 
    
    
    my_params = {
        'q':"",#q is query
        'from-date': "", #begin-date
        'to-date': "", #finish-date
        'order-by':"oldest", 
        'show-fields': "all",
        'page-size':10 ,
        'api-key': API_KEY,
        'tag':"",
        'section':"",
        'lang':"",
    }
    
    my_params['from-date'] = "2019-08-19"
    my_params['to-date'] = "2019-08-19"
    my_params['tag'] = "politics/politics"
    my_params['q'] = "bank"
    my_params['section'] = "politics"
    my_params['lang'] = "en"
    
    body_list=[] 
    topic=[] 
    startrow=None 
    current_page = 1
    total_pages = 1
    
    while current_page <= total_pages:
                print("...page", current_page)
                my_params['page'] = current_page
                resp = requests.get(API_ENDPOINT, my_params)
                print(resp.url)
                print("\n")
                data = resp.json()
                current_page += 1
                total_pages=data['response']['pages']

                results=len(data["response"]["results"])
                
                for i in range(0,results):
                    d=data["response"]["results"][i]["fields"]["bodyText"]
                    d.replace("\n","")
                    dd=html2text.html2text(d) 
                    dd = re.sub(r"\(https\:\S+|\s+\)", "", dd)
                    body_list.append(dd)
                    topic.append(my_params.get("q")) 
                    
                list_of_tuples = list(zip(body_list, topic))  
                
                df = pd.DataFrame(list_of_tuples,index=None, columns=None)
                
                filename = r"D:\guardian.xls"
            
                df.to_excel("q.xlsx",index=False,header=False) # it overwrites the excel file 
                """
                # it appends the dataframe df to the excel
                writer = pd.ExcelWriter(filename, engine='openpyxl')
                writer.book = load_workbook(filename)
                if  "Sheet1" in writer.book.sheetnames:
                    startrow = writer.book["Sheet1"].max_row
                    
                writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
                
                if startrow is None:
                    startrow = 0

                df.to_excel(writer, "Sheet1", startrow=startrow,index=False,header=False)
                writer.save()
                writer.close()
    """           

schedule.every().day.at("17:21").do(guardian) #specify the time to automatically run the api

while True:
    schedule.run_pending()
    time.sleep(1) 

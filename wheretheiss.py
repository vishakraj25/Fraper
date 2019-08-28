"""
Author : Vishak Raj
E-mail ID : vishak.shanmu@gmail.com

Gives and save the wheretheiss details and have option to get details by mentioning date and the interval between date
And the wheretheiss.at may have limits 

"""
#wheretheiss.at
import json
import requests
import time
import datetime as dt
import html2text,pandas as pd
import csv,xlsxwriter
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import re
import schedule
import time

#Refer the doc for more information - https://wheretheiss.at/w/developer

def iss():
    API_POSITION_ENDPOINT = 'https://api.wheretheiss.at/v1/satellites/25544/positions'

    my_params = {
    'timestamps':"",
    'units':"",
    }

    begin_date="2014-01-01"
    end_date="2014-02-11"
    date_format = "%Y-%m-%d"
    a = dt.datetime.strptime(begin_date, "%Y-%m-%d")
    b = dt.datetime.strptime(end_date, "%Y-%m-%d")
    delta = b - a
    total_days = delta.days

    timestamps=[]
    t1 = dt.datetime.strptime(begin_date, "%Y-%m-%d")
    timestamps.append(dt.datetime.timestamp(t1))

    while(t1<=b):
        t1+= dt.timedelta(hours=5)#,seconds=s, minutes=m,days=1, ) add data for how much difference need in between two dates i.e., begin and end date
        if t1<=b:      
            timestamps.append(dt.datetime.timestamp(t1))

    composite_list = [timestamps[x:x+10] for x in range(0, len(timestamps),10)]
    for i in composite_list:
        timestamp_string = ','.join(str(e) for e in i)
        my_params['timestamps'] = timestamp_string
        my_params['units'] = "kilometers"

        resp = requests.get(API_POSITION_ENDPOINT, my_params)
        print(resp.url)
        data=resp.json()
        results=len(data)
        for i in range(0,results):
            d=data[i]
            values=list(d.values())
            df = pd.DataFrame(values,columns=None)
            df = df.transpose()

            filename = r"D:\iss.xlsx"
            df.to_excel(filename,index=False,header=False) #overwriteing the data in the excel file 

            """
                # for appending the data in same excel file 

            writer = pd.ExcelWriter(filename, engine='openpyxl')
            writer.book = load_workbook(filename)
            if  "Sheet1" in writer.book.sheetnames:
                startrow = writer.book["Sheet1"].max_row

            writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

            if startrow is None:
                startrow = 0

            df.to_excel(writer, "Sheet1", startrow=startrow,index=False,header=False)

            writer.save()
            writer.close()

            """


schedule.every().day.at("17:21").do(iss) #specify the time to automatically run the api

while True:
    schedule.run_pending()
    time.sleep(1) 
